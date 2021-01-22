









import pprint



from openpyxl import load_workbook
workbook=load_workbook('case_v1.xlsx')#加载文件得到一个wook_book对象

sheet=workbook['login']#获取表单
data=list(sheet.rows)#获取所有行并转换成列表
title=[]
reslut=[]
for k in (data[0]):#获取标题
    title.append(k.value)
for i in (data[1:]):#获取所有内容
   dic={}
   for k,v in enumerate(i):#获取每行的索引及值
       dic[title[k]]=v.value
   reslut.append(dic)
print(reslut)
# sheet.cell(1,6).value="测试组"#指定单元格并写入的内容
# workbook.save('case_v1.xlsx')#保存
# workbook.close()#关闭
# sheetnames=workbook.sheetnames
# # print(sheetnames)

# data=sheet.cell(1,2)#第一行第二列的单元格
# print(data)
# print(data.value)#单元格的值
# data=sheet.rows#获取所有行
# print(list(data))